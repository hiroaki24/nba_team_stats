from nba_api.stats.endpoints import leaguedashteamstats
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

team_abbr = input('チーム略称を入力してください: ').upper()

team_map = {
    'ATL':'Atlanta Hawks','BOS':'Boston Celtics','BKN':'Brooklyn Nets',
    'CHA':'Charlotte Hornets','CHI':'Chicago Bulls','CLE':'Cleveland Cavaliers',
    'DAL':'Dallas Mavericks','DEN':'Denver Nuggets','DET':'Detroit Pistons',
    'GSW':'Golden State Warriors','HOU':'Houston Rockets','IND':'Indiana Pacers',
    'LAC':'LA Clippers','LAL':'Los Angeles Lakers','MEM':'Memphis Grizzlies',
    'MIA':'Miami Heat','MIL':'Milwaukee Bucks','MIN':'Minnesota Timberwolves',
    'NOP':'New Orleans Pelicans','NYK':'New York Knicks','OKC':'Oklahoma City Thunder',
    'ORL':'Orlando Magic','PHI':'Philadelphia 76ers','PHX':'Phoenix Suns',
    'POR':'Portland Trail Blazers','SAC':'Sacramento Kings','SAS':'San Antonio Spurs',
    'TOR':'Toronto Raptors','UTA':'Utah Jazz','WAS':'Washington Wizards'
}

if team_abbr not in team_map:
    print('無効な略称です')
    exit()

team_name = team_map[team_abbr]

df = leaguedashteamstats.LeagueDashTeamStats(
    season='2025-26',
    season_type_all_star='Regular Season',
    per_mode_detailed='PerGame',
    measure_type_detailed_defense='Advanced'
).get_data_frames()[0]

out = df[['TEAM_NAME','W','L','OFF_RATING','DEF_RATING']].copy()
out['W_RANK'] = out['W'].rank(ascending=False, method='min').astype(int)
out['L_RANK'] = out['L'].rank(ascending=True, method='min').astype(int)
out['OFF_RATING_RANK'] = out['OFF_RATING'].rank(ascending=False, method='min').astype(int)
out['DEF_RATING_RANK'] = out['DEF_RATING'].rank(ascending=True, method='min').astype(int)

team_out = out[out['TEAM_NAME'] == team_name].copy()

file_name = f'{team_abbr}_team_stats.xlsx'
team_out.to_excel(file_name, index=False)

wb = load_workbook(file_name)
ws = wb.active
red = PatternFill(fill_type='solid', fgColor='FFC7CE')
blue = PatternFill(fill_type='solid', fgColor='CFE2F3')

for col in [6, 7, 8, 9]:
    v = ws.cell(row=2, column=col).value
    if v <= 10:
        ws.cell(row=2, column=col).fill = red
    elif v >= 21:
        ws.cell(row=2, column=col).fill = blue

wb.save(file_name)
print(f'{file_name} を保存しました')
print(team_out)